import pandas as pd
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from openpyxl.utils import get_column_letter \



# def mjm():
#     """
#     门静脉特殊处理
#     :return:
#     """
#     n = 1
#     mjm_data = pd.read_excel('腹部脏器统计_20200823.xlsx', header=1, sheetname='门静脉')
#     pjm = mjm_data.iloc[:, 1:6]
#     mjm = mjm_data.iloc[:, 6:11]
#     cxmsjm = mjm_data.iloc[:, 11:16]
#     datas = [pjm, mjm, cxmsjm]
#     for data in datas:
#         # 遍历所有维度
#         for j in data.columns:
#             index = list(data.columns).index(j)
#             data.insert(index + 1, j + '-test', np.nan)
#
#             current_column = data.loc[:, j]
#             ymean = np.mean(current_column)
#             ystd = np.std(current_column)
#             floor = ymean - (n * ystd)
#             upper = ymean + (n * ystd)
#             y = current_column
#             for i in range(len(data)):
#                 if y[i] < floor:
#                     data.loc[i, j + '-test'] = 'Shrunken'
#                 if y[i] > upper:
#                     data.loc[i, j + '-test'] = 'Enlarged'
#     all_data = pd.concat([data for data in datas], axis=1)
#     columns_list = []
#     for i in all_data.columns:
#         for j in ['.1', '.2', '.3', '.4']:
#             if j in i:
#                 i = i.replace(j, '')
#         columns_list.append(i)
#     all_data.columns = columns_list
#     all_data.insert(0, 'A号', mjm_data.loc[:, 'A号'])
#     output_file = r'C:\Users\SMIT\Desktop\腹部脏器统计_20200823.xlsx'
#     if os.path.exists(output_file):
#         writer = pd.ExcelWriter(output_file, engine='openpyxl')
#         book = load_workbook(output_file)
#         writer.book = book
#         all_data.to_excel(writer, sheet_name='门静脉' + '-', index=False)
#         writer.save()
#
#
# mjm()
# output_file = r'C:\Users\SMIT\Desktop\腹部脏器统计_20200823.xlsx'
#
#
# def operation_excel():
#     output_file = r'C:\Users\SMIT\Desktop\腹部脏器统计_20200823.xlsx'
#     wb_new = Workbook()
#     wb_old = load_workbook(output_file)
#     thin = Side(border_style="thin", color="000000")
#     double = Side(border_style="thin", color="000000")
#     fille = PatternFill('solid', fgColor='D2E9FF')
#     fillr = PatternFill('solid', fgColor='FFE4CA')
#     fonte = Font(name=u'等线', bold=True, italic=False, size=11)
#     # 遍历源excel中的sheet
#     for sheet in wb_old:
#         if sheet == '门静脉-':
#             wb_new = Workbook()
#             wb_old = load_workbook(output_file)
#             thin = Side(border_style="thin", color="000000")
#             double = Side(border_style="thin", color="000000")
#             fille = PatternFill('solid', fgColor='D2E9FF')
#             fillr = PatternFill('solid', fgColor='FFE4CA')
#             fonte = Font(name=u'等线', bold=True, italic=False, size=11)
#             # 遍历源excel中的sheet
#             sheet = wb_old['门静脉-']
#
#             # 新建与源sheet同名sheet
#             ws_new = wb_new.create_sheet(sheet.title.replace('-', ''))
#             # 01写入表头
#             A1 = sheet.max_row
#             ws_new.append([A1])
#             # 注意先保存文件，否则文件不存在
#             new_xlsx_path = output_file
#             wb_new.save(new_xlsx_path)
#
#             # 02复制原sheet内容
#             ws_old = wb_old[sheet.title]
#             ws_new = wb_new[ws_new.title]
#             # 两个for循环遍历整个源sheet的单元格内容
#             for i, row in enumerate(ws_old.iter_rows()):
#                 for j, cell in enumerate(row):
#                     # 因为插入一行表头，所以行数整体下移一行写入
#                     ws_new.cell(row=i + 2, column=j + 1, value=cell.value)
#             ws_new.merge_cells('B1:K1')
#             merge_cell = ws_new['B1']
#             merge_cell.value = '脾静脉-label__1008-86'
#             merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#             merge_cell.fill = fillr
#             merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#             for item in ws_new['B2':'K{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#                 for cell in item:  # cell表示每一行的每一个单元格对象
#                     cell.fill = fille
#                     cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#
#             ws_new.merge_cells('L1:U1')
#             merge_cell = ws_new['L1']
#             merge_cell.value = '门静脉-label__1208'
#             merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#             merge_cell.fill = fillr
#             merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#             for item in ws_new['L2':'U{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#                 for cell in item:  # cell表示每一行的每一个单元格对象
#                     cell.fill = fillr
#                     cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#             ws_new.merge_cells('V1:AE1')
#             merge_cell = ws_new['V1']
#             merge_cell.value = '肠系膜上静脉-label__1212'
#             merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#             merge_cell.fill = fille
#             merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#             for item in ws_new['V2':'AE{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#                 for cell in item:  # cell表示每一行的每一个单元格对象
#                     cell.fill = fille
#                     cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#             # 字体设置
#             for item in ws_new['A1':'AE2']:
#                 for cell in item:
#                     cell.font = fonte
#                     cell.alignment = Alignment(horizontal='center', vertical='center')
#             # 自适应列宽
#             col_width = []
#             i = 0
#             # 每列
#             for col in sheet.columns:
#                 # 每行
#                 for j in range(len(col)):
#                     if j == 0:
#                         # 数组增加一个元素
#                         col_width.append(len(str(col[j].value)))
#                     else:
#                         # 获得每列中的内容的最大宽度
#                         if col_width[i] < len(str(col[j].value)):
#                             col_width[i] = len(str(col[j].value))
#                 i = i + 1
#
#             # 设置列宽
#             for i in range(len(col_width)):
#                 # 根据列的数字返回字母
#                 col_letter = get_column_letter(i + 1)
#                 # 当宽度大于100，宽度设置为100
#                 if col_width[i] > 100:
#                     ws_new.column_dimensions[col_letter].width = 100
#                 # 只有当宽度大于10，才设置列宽
#                 elif col_width[i] > 1:
#                     ws_new.column_dimensions[col_letter].width = col_width[i] + 2
#                     # 删除新建excel中多余的'sheet'页
#
#             wb_new.save(new_xlsx_path)
#             continue
#         # 新建与源sheet同名sheet
#         ws_new = wb_new.create_sheet(sheet.title.replace('-', ''))
#         # 01写入表头
#         A1 = sheet.max_row
#         ws_new.append([A1])
#         # 注意先保存文件，否则文件不存在
#         new_xlsx_path = output_file
#         wb_new.save(new_xlsx_path)
#
#         # 02复制原sheet内容
#         ws_old = wb_old[sheet.title]
#         ws_new = wb_new[ws_new.title]
#         # 两个for循环遍历整个源sheet的单元格内容
#         for i, row in enumerate(ws_old.iter_rows()):
#             for j, cell in enumerate(row):
#                 # 因为插入一行表头，所以行数整体下移一行写入
#                 ws_new.cell(row=i + 2, column=j + 1, value=cell.value)
#         ws_new.merge_cells('B1:M1')
#         merge_cell = ws_new['B1']
#         merge_cell.value = '径线X'
#         merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#         merge_cell.fill = fille
#         merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#         for item in ws_new['B2':'M{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#             for cell in item:  # cell表示每一行的每一个单元格对象
#                 cell.fill = fille
#                 cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#
#         ws_new.merge_cells('N1:Y1')
#         merge_cell = ws_new['N1']
#         merge_cell.value = '径线Y'
#         merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#         merge_cell.fill = fillr
#         merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#         for item in ws_new['N2':'Y{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#             for cell in item:  # cell表示每一行的每一个单元格对象
#                 cell.fill = fillr
#                 cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#
#         ws_new.merge_cells('Z1:AK1')
#         merge_cell = ws_new['Z1']
#         merge_cell.value = '径线Z'
#         merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#         merge_cell.fill = fille
#         merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#         for item in ws_new['Z2':'AK{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#             for cell in item:  # cell表示每一行的每一个单元格对象
#                 cell.fill = fille
#                 cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#
#         ws_new.merge_cells('AL1:AW1')
#         merge_cell = ws_new['AL1']
#         merge_cell.value = '体积'
#         merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#         merge_cell.fill = fillr
#         merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#         for item in ws_new['AL2':'AW{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#             for cell in item:  # cell表示每一行的每一个单元格对象
#                 cell.fill = fillr
#                 cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#
#         ws_new.merge_cells('AX1:BI1')
#         merge_cell = ws_new['AX1']
#         merge_cell.value = '平均CT值'
#         merge_cell.alignment = Alignment(horizontal='center', vertical='center')
#         merge_cell.fill = fille
#         merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#         for item in ws_new['AX2':'BI{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
#             for cell in item:  # cell表示每一行的每一个单元格对象
#                 cell.fill = fille
#                 cell.border = Border(top=double, left=thin, right=thin, bottom=double)
#
#         # 字体设置
#         for item in ws_new['A1':'BI2']:
#             for cell in item:
#                 cell.font = fonte
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#         # 自适应列宽
#         col_width = []
#         i = 0
#         # 每列
#         for col in sheet.columns:
#             # 每行
#             for j in range(len(col)):
#                 if j == 0:
#                     # 数组增加一个元素
#                     col_width.append(len(str(col[j].value)))
#                 else:
#                     # 获得每列中的内容的最大宽度
#                     if col_width[i] < len(str(col[j].value)):
#                         col_width[i] = len(str(col[j].value))
#             i = i + 1
#
#         # 设置列宽
#         for i in range(len(col_width)):
#             # 根据列的数字返回字母
#             col_letter = get_column_letter(i + 1)
#             # 当宽度大于100，宽度设置为100
#             if col_width[i] > 100:
#                 ws_new.column_dimensions[col_letter].width = 100
#             # 只有当宽度大于10，才设置列宽
#             elif col_width[i] > 1:
#                 ws_new.column_dimensions[col_letter].width = col_width[i] + 2
#     # 删除新建excel中多余的'sheet'页
#     wb_new.remove_sheet(wb_new.get_sheet_by_name('Sheet'))
#     # 保存新excel
#     wb_new.save(new_xlsx_path)
#
# operation_excel()

output_file = r'C:\Users\SMIT\Desktop\腹部脏器统计_20200823.xlsx'
wb_new = Workbook()
wb_old = load_workbook(output_file)
thin = Side(border_style="thin", color="000000")
double = Side(border_style="thin", color="000000")
fille = PatternFill('solid', fgColor='D2E9FF')
fillr = PatternFill('solid', fgColor='FFE4CA')
fonte = Font(name=u'等线', bold=True, italic=False, size=11)
# 遍历源excel中的sheet
for sheet in wb_old:
    # 新建与源sheet同名sheet
    ws_new = wb_new.create_sheet(sheet.title.replace('-', ''))
    # 01写入表头
    A1 = sheet.max_row
    ws_new.append([A1])
    # 注意先保存文件，否则文件不存在
    new_xlsx_path = output_file
    wb_new.save(new_xlsx_path)

    # 02复制原sheet内容
    ws_old = wb_old[sheet.title]
    ws_new = wb_new[ws_new.title]
    # 两个for循环遍历整个源sheet的单元格内容
    for i, row in enumerate(ws_old.iter_rows()):
        for j, cell in enumerate(row):
            # 因为插入一行表头，所以行数整体下移一行写入
            ws_new.cell(row=i + 2, column=j + 1, value=cell.value)
    if sheet.title == '门静脉-':
        ws_new.merge_cells('B1:K1')
        merge_cell = ws_new['B1']
        merge_cell.value = '脾静脉-label__1008-86'
        merge_cell.alignment = Alignment(horizontal='center', vertical='center')
        merge_cell.fill = fille
        merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        for item in ws_new['B2':'K{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
            for cell in item:  # cell表示每一行的每一个单元格对象
                cell.fill = fille
                cell.border = Border(top=double, left=thin, right=thin, bottom=double)

        ws_new.merge_cells('L1:U1')
        merge_cell = ws_new['L1']
        merge_cell.value = '门静脉-label__1208'
        merge_cell.alignment = Alignment(horizontal='center', vertical='center')
        merge_cell.fill = fillr
        merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        for item in ws_new['L2':'U{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
            for cell in item:  # cell表示每一行的每一个单元格对象
                cell.fill = fillr
                cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        ws_new.merge_cells('V1:AE1')
        merge_cell = ws_new['V1']
        merge_cell.value = '肠系膜上静脉-label__1212'
        merge_cell.alignment = Alignment(horizontal='center', vertical='center')
        merge_cell.fill = fille
        merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        for item in ws_new['V2':'AE{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
            for cell in item:  # cell表示每一行的每一个单元格对象
                cell.fill = fille
                cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        # 字体设置
        for item in ws_new['A1':'AE2']:
            for cell in item:
                cell.font = fonte
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # 自适应列宽
        col_width = []
        i = 0
        # 每列
        for col in sheet.columns:
            # 每行
            for j in range(len(col)):
                if j == 0:
                    # 数组增加一个元素
                    col_width.append(len(str(col[j].value)))
                else:
                    # 获得每列中的内容的最大宽度
                    if col_width[i] < len(str(col[j].value)):
                        col_width[i] = len(str(col[j].value))
            i = i + 1
        # 设置列宽
        for i in range(len(col_width)):
            # 根据列的数字返回字母
            col_letter = get_column_letter(i + 1)
            # 当宽度大于100，宽度设置为100
            if col_width[i] > 100:
                ws_new.column_dimensions[col_letter].width = 100
            # 只有当宽度大于10，才设置列宽
            elif col_width[i] > 1:
                ws_new.column_dimensions[col_letter].width = col_width[i] + 2
        continue
    ws_new.merge_cells('B1:M1')
    merge_cell = ws_new['B1']
    merge_cell.value = '径线X'
    merge_cell.alignment = Alignment(horizontal='center', vertical='center')
    merge_cell.fill = fille
    merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    for item in ws_new['B2':'M{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
        for cell in item:  # cell表示每一行的每一个单元格对象
            cell.fill = fille
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)

    ws_new.merge_cells('N1:Y1')
    merge_cell = ws_new['N1']
    merge_cell.value = '径线Y'
    merge_cell.alignment = Alignment(horizontal='center', vertical='center')
    merge_cell.fill = fillr
    merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    for item in ws_new['N2':'Y{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
        for cell in item:  # cell表示每一行的每一个单元格对象
            cell.fill = fillr
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)

    ws_new.merge_cells('Z1:AK1')
    merge_cell = ws_new['Z1']
    merge_cell.value = '径线Z'
    merge_cell.alignment = Alignment(horizontal='center', vertical='center')
    merge_cell.fill = fille
    merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    for item in ws_new['Z2':'AK{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
        for cell in item:  # cell表示每一行的每一个单元格对象
            cell.fill = fille
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)

    ws_new.merge_cells('AL1:AW1')
    merge_cell = ws_new['AL1']
    merge_cell.value = '体积'
    merge_cell.alignment = Alignment(horizontal='center', vertical='center')
    merge_cell.fill = fillr
    merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    for item in ws_new['AL2':'AW{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
        for cell in item:  # cell表示每一行的每一个单元格对象
            cell.fill = fillr
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)

    ws_new.merge_cells('AX1:BI1')
    merge_cell = ws_new['AX1']
    merge_cell.value = '平均CT值'
    merge_cell.alignment = Alignment(horizontal='center', vertical='center')
    merge_cell.fill = fille
    merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    for item in ws_new['AX2':'BI{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
        for cell in item:  # cell表示每一行的每一个单元格对象
            cell.fill = fille
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)

    # 字体设置
    for item in ws_new['A1':'BI2']:
        for cell in item:
            cell.font = fonte
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 自适应列宽
    col_width = []
    i = 0
    # 每列
    for col in sheet.columns:
        # 每行
        for j in range(len(col)):
            if j == 0:
                # 数组增加一个元素
                col_width.append(len(str(col[j].value)))
            else:
                # 获得每列中的内容的最大宽度
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        i = i + 1

    # 设置列宽
    for i in range(len(col_width)):
        # 根据列的数字返回字母
        col_letter = get_column_letter(i + 1)
        # 当宽度大于100，宽度设置为100
        if col_width[i] > 100:
            ws_new.column_dimensions[col_letter].width = 100
        # 只有当宽度大于10，才设置列宽
        elif col_width[i] > 1:
            ws_new.column_dimensions[col_letter].width = col_width[i] + 2
# 删除新建excel中多余的'sheet'页
wb_new.remove_sheet(wb_new.get_sheet_by_name('Sheet'))
# 保存新excel
wb_new.save(new_xlsx_path)