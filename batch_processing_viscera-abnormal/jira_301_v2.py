import pandas as pd
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from openpyxl.utils import get_column_letter
import xlrd


class VisveraAbnormal:

    def __init__(self, input_file, n, output_path=None):
        if os.path.isfile(input_file):
            pass
        else:
            raise Exception('输入不是文件！')
        self._file = input_file
        self._n = n
        if output_path:
            self._output_path = output_path
        else:
            p, f = os.path.split(input_file)
            if os.path.exists(os.path.join(p, 'new_zq')):
                pass
            else:
                os.makedirs(os.path.join(p, 'new_zq'))
            self._output_path = os.path.join(p, 'new_zq')

    def get_sheet(self):
        """
        获取excel文件所有sheet
        :return:
        """
        sheet_name_list = []
        file = pd.ExcelFile(self._file)
        for i in file.sheet_names[0:5]:
            sheet_name_list.append(i)
        sheet_name_list.append('门静脉')
        return sheet_name_list

    def build_dataframe(self, sheetname):
        """
        将excel中每个sheet的数据都遍历执行，构建dataframe
        :return:
        """
        # excel所有数据
        all_data = pd.read_excel(self._file, header=1, sheetname=sheetname)
        self._all_data = all_data
        self._sheetname = sheetname
        radial_line_x = all_data.loc[:, ['NoC', 'AP', 'PVP', 'DP-early', 'DP-item', 'DP-late', 'Ave', 'AveStd', 'Std',  'three-qx']]
        radial_line_y = all_data.loc[:, ['NoC.1', 'AP.1', 'PVP.1', 'DP-early.1', 'DP-item.1', 'DP-late.1', 'Ave', 'AveStd', 'Std', 'three-qx']]
        radial_line_z = all_data.loc[:, ['NoC.2', 'AP.2', 'PVP.2', 'DP-early.2', 'DP-item.2', 'DP-late.2', 'Ave', 'AveStd', 'Std', 'three-qx']]
        volume = all_data.loc[:, ['NoC.3', 'AP.3', 'PVP.3', 'DP-early.3', 'DP-item.3', 'DP-late.3', 'Ave', 'AveStd', 'Std', 'three-qx']]
        ct_avg = all_data.loc[:, ['NoC.4', 'AP.4', 'PVP.4', 'DP-early.4', 'DP-item.4', 'DP-late.4', 'Ave', 'AveStd', 'Std', 'three-qx']]
        datas = [radial_line_x, radial_line_y, radial_line_z, volume, ct_avg]
        values = radial_line_x.loc[0, 'three-qx']
        for data in datas:
            tmp = data.fillna(0)
            data.loc[:, 'three-qx'] = tmp.iloc[:, 2] + tmp.iloc[:, 3] + tmp.iloc[:, 4] + tmp.iloc[:, 5]
            # 构建新的列，用来检测每个值是否是异常值
            for i in data.columns[:-1]:
                index = list(data.columns[:-1]).index(i)
                data.insert(index + 1, i + '-test', values)
        return datas

    def save_excel(self, datas, output_path):
        all_data = pd.concat([data.iloc[:, :-1] for data in datas], axis=1)
        all_data = all_data.round(2)
        all_data = all_data.fillna('N/A')
        columns_list = []
        for i in all_data.columns:
            for j in ['.1', '.2', '.3', '.4']:
                if j in i:
                    i = i.replace(j, '')
            columns_list.append(i)
        all_data.columns = columns_list
        all_data.insert(0, 'A号', self._all_data.loc[:, 'A号'])
        p, f = os.path.split(self._file)
        output_file = os.path.join(output_path, f)
        self._output_file = output_file
        if os.path.exists(output_file):
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            book = load_workbook(output_file)
            writer.book = book
            all_data.to_excel(writer, sheet_name=self._sheetname + '-', index=False)
            writer.save()
        else:
            writer = pd.ExcelWriter(output_file)
            all_data.to_excel(writer, sheet_name=self._sheetname + '-', index=False)
            writer.save()

    def col_analyze(self, datas):
        """
        :param    构建的DataFrame数据列表

                   NoC         AP        PVP   DP-early    DP-item
        0      30.242556  39.965484        NaN        NaN  32.333421
        1      43.164668  43.310709        NaN  46.294548        NaN
        2      43.922559  42.768043        NaN  44.396025        NaN
        3      45.331145  37.447831  21.212967        NaN        NaN
        4      43.985022  40.447139        NaN  33.659616        NaN

        :return:    在构建好的DataFrame基础上，每列后增加一列备注列进行检测异常值

                    NoC NoC-test         AP AP-test      PVP PVP-test   DP-early DP-early-test    DP-item   DP-item-test  DP-late  DP-late-test     Ave    Ave-test ...  three-qx
        0    30.242556    NaN    39.965484    NaN         NaN   NaN           NaN      NaN    32.333421         NaN          NaN     NaN         34.180487    NaN   ...  32.333421
        1    43.164668    NaN    43.310709    NaN         NaN   NaN     46.294548      NaN          NaN         NaN          NaN     NaN         44.256642    NaN   ...  46.294548
        2    43.922559    NaN    42.768043    NaN         NaN   NaN     44.396025      NaN          NaN         NaN          NaN     NaN         43.695542    NaN   ...  44.396025
        3    45.331145    NaN    37.447831    NaN   21.212967   NaN           NaN      NaN          NaN         NaN          NaN     NaN         34.663981    NaN   ...  21.212967
        4    43.985022    NaN    40.447139    NaN         NaN   NaN     33.659616      NaN          NaN         NaN          NaN     NaN         39.363926    NaN   ...  33.659616
        5    53.188478    NaN    41.666311    NaN         NaN   NaN     44.803421      NaN          NaN         NaN          NaN     NaN         46.552737    NaN   ...  44.803421
        6    23.753227    NaN    42.652325    NaN         NaN   NaN     22.783115      NaN          NaN         NaN          NaN     NaN         29.729556    NaN   ...  22.783115
        """
        # 列方向
        for data in datas:
            tmps = data.iloc[:, [0, 2, -1]]
            #     异常值检测3sigma原则
            for i in range(len(tmps)):
                demo = tmps.iloc[i, :]
                # print(demo)
                # 均值
                ymean = np.mean(demo)
                # 标准差
                ystd = np.std(demo)
                # 平均离差
                num = 0
                for j in tmps.columns:
                    num += abs(data.loc[i, j] - ymean)
                ymena_std = num/len(demo)
                data.loc[i, 'Ave'] = ymean
                data.loc[i, 'Std'] = ystd
                data.loc[i, 'AveStd'] = ymena_std
        self.row_analyze(datas)

    def row_analyze(self, datas):
        n = self._n
        # 行方向
        for data in datas:
            # 获取列维度
            tmps = data.iloc[:, [0, 2, -7, -5, -3, -1]]
            # print(tmps)
            #     异常值检测3sigma原则
            for i in tmps.columns:
                # print(i)
                demo = tmps.loc[:, i]
                ymean = np.mean(demo)
                ystd = np.std(demo)
                floor = ymean - (n * ystd)
                upper = ymean + (n * ystd)
                for l in range(len(data)):
                    if i == 'three-qx':
                        for z in data.iloc[:, [4, 6, 8, 10]]:
                            # print(z)
                            k = data.loc[l, z]
                            if (k < floor):
                                data.loc[l, z + '-test'] = 'Shrunken'
                            elif (k > upper):
                                data.loc[l, z + '-test'] = 'Enlarged'
                            elif (k > floor) and (k < upper):
                                data.loc[l, z + '-test'] = 'Normal'
                    else:
                        y = data.loc[l, i]
                        if (y < floor):
                                data.loc[l, i + '-test'] = 'Shrunken'
                        elif (y > upper):
                                data.loc[l, i + '-test'] = 'Enlarged'
                        elif (y > floor) and (y < upper):
                            data.loc[l, i + '-test'] = 'Normal'
        self.save_excel(datas, self._output_path)

    def mjm(self):
        """
        门静脉特殊处理
        :return:
        """
        n = 1
        mjm_data = pd.read_excel('腹部脏器统计_20200823.xlsx', header=1, sheetname='门静脉')
        pjm = mjm_data.iloc[:, 1:6]
        mjm = mjm_data.iloc[:, 6:11]
        cxmsjm = mjm_data.iloc[:, 11:16]
        datas = [pjm, mjm, cxmsjm]
        for data in datas:
            # 遍历所有维度
            for j in data.columns:
                index = list(data.columns).index(j)
                data.insert(index + 1, j + '-test', np.nan)
                current_column = data.loc[:, j]
                ymean = np.mean(current_column)
                ystd = np.std(current_column)
                floor = ymean - (n * ystd)
                upper = ymean + (n * ystd)
                y = current_column
                for i in range(len(data)):
                    if y[i] < floor:
                        data.loc[i, j + '-test'] = 'Shrunken'
                    elif y[i] > upper:
                        data.loc[i, j + '-test'] = 'Enlarged'
                    elif y[i] > floor and y[i] < upper:
                        data.loc[i, j + '-test'] = 'Normal'
        all_data = pd.concat([data for data in datas], axis=1)
        all_data = all_data.round(2)
        all_data = all_data.fillna('N/A')
        columns_list = []
        for i in all_data.columns:
            for j in ['.1', '.2', '.3', '.4']:
                if j in i:
                    i = i.replace(j, '')
            columns_list.append(i)
        all_data.columns = columns_list
        all_data.insert(0, 'A号', mjm_data.loc[:, 'A号'])
        p, f = os.path.split(self._file)
        output_file = os.path.join(self._output_path, f)
        if os.path.exists(output_file):
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            book = load_workbook(output_file)
            writer.book = book
            all_data.to_excel(writer, sheet_name='门静脉' + '-', index=False)
            writer.save()

    def operation_excel(self):
        output_file = self._output_file
        wb_new = Workbook()
        wb_old = load_workbook(output_file)
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="thin", color="000000")
        fille = PatternFill('solid', fgColor='D2E9FF')
        fillr = PatternFill('solid', fgColor='FFE4CA')
        fonte = Font(name=u'等线', bold=True, italic=False, size=11)

        # 遍历源excel中的sheet
        for sheet in wb_old:
            # 新建与源sheet同名sheet
            ws_new = wb_new.create_sheet(sheet.title.replace('-', ''))
            # 01写入表头
            A1 = sheet.max_row-1
            ws_new.append([A1])
            # 注意先保存文件，否则文件不存在
            new_xlsx_path = output_file
            wb_new.save(new_xlsx_path)

            # 02复制原sheet内容
            ws_old = wb_old[sheet.title]
            ws_new = wb_new[ws_new.title]
            # 两个for循环遍历整个源sheet的单元格内容
            for i, row in enumerate(ws_old.iter_rows()):
                for j, cell in enumerate(row):
                    # 因为插入一行表头，所以行数整体下移一行写入
                    ws_new.cell(row=i + 2, column=j + 1, value=cell.value)
            if sheet.title == '门静脉-':
                ws_new.merge_cells('B1:K1')
                merge_cell = ws_new['B1']
                merge_cell.value = '脾静脉-label__1008-86'
                merge_cell.alignment = Alignment(horizontal='center', vertical='center')
                merge_cell.fill = fille
                merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
                for item in ws_new['B2':'K{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                    for cell in item:  # cell表示每一行的每一个单元格对象
                        cell.fill = fille
                        cell.border = Border(top=double, left=thin, right=thin, bottom=double)

                ws_new.merge_cells('L1:U1')
                merge_cell = ws_new['L1']
                merge_cell.value = '门静脉-label__1208'
                merge_cell.alignment = Alignment(horizontal='center', vertical='center')
                merge_cell.fill = fillr
                merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
                for item in ws_new['L2':'U{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                    for cell in item:  # cell表示每一行的每一个单元格对象
                        cell.fill = fillr
                        cell.border = Border(top=double, left=thin, right=thin, bottom=double)
                ws_new.merge_cells('V1:AE1')
                merge_cell = ws_new['V1']
                merge_cell.value = '肠系膜上静脉-label__1212'
                merge_cell.alignment = Alignment(horizontal='center', vertical='center')
                merge_cell.fill = fille
                merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
                for item in ws_new['V2':'AE{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                    for cell in item:  # cell表示每一行的每一个单元格对象
                        cell.fill = fille
                        cell.border = Border(top=double, left=thin, right=thin, bottom=double)
                # 字体设置
                for item in ws_new['A2':'AE{}'.format(sheet.max_row + 1)]:
                    for cell in item:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                for item in ws_new['A1':'AE2']:
                    for cell in item:
                        cell.font = fonte
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                # 自适应列宽
                col_width = []
                i = 0
                # 每列
                for col in sheet.columns:
                    # 每行
                    for j in range(len(col)):
                        if j == 0:
                            # 数组增加一个元素
                            col_width.append(len(str(col[j].value)))
                        else:
                            # 获得每列中的内容的最大宽度
                            if col_width[i] < len(str(col[j].value)):
                                col_width[i] = len(str(col[j].value))
                    i = i + 1

                # 设置列宽
                for i in range(len(col_width)):
                    # 根据列的数字返回字母
                    col_letter = get_column_letter(i + 1)
                    # 当宽度大于100，宽度设置为100
                    if col_width[i] > 100:
                        ws_new.column_dimensions[col_letter].width = 100
                    # 只有当宽度大于10，才设置列宽
                    elif col_width[i] > 1:
                        ws_new.column_dimensions[col_letter].width = col_width[i] + 2
                continue
            ws_new.merge_cells('B1:S1')
            merge_cell = ws_new['B1']
            merge_cell.value = '径线X'
            merge_cell.alignment = Alignment(horizontal='center', vertical='center')
            merge_cell.fill = fille
            merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            for item in ws_new['B2':'S{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                for cell in item:  # cell表示每一行的每一个单元格对象
                    cell.fill = fille
                    cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            ws_new.merge_cells('T1:AK1')
            merge_cell = ws_new['T1']
            merge_cell.value = '径线Y'
            merge_cell.alignment = Alignment(horizontal='center', vertical='center')
            merge_cell.fill = fillr
            merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            for item in ws_new['T2':'AK{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                for cell in item:  # cell表示每一行的每一个单元格对象
                    cell.fill = fillr
                    cell.border = Border(top=double, left=thin, right=thin, bottom=double)

            ws_new.merge_cells('AL1:BC1')
            merge_cell = ws_new['AL1']
            merge_cell.value = '径线Z'
            merge_cell.alignment = Alignment(horizontal='center', vertical='center')
            merge_cell.fill = fille
            merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            for item in ws_new['AL2':'BC{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                for cell in item:  # cell表示每一行的每一个单元格对象
                    cell.fill = fille
                    cell.border = Border(top=double, left=thin, right=thin, bottom=double)

            ws_new.merge_cells('BD1:BU1')
            merge_cell = ws_new['BD1']
            merge_cell.value = '体积'
            merge_cell.alignment = Alignment(horizontal='center', vertical='center')
            merge_cell.fill = fillr
            merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            for item in ws_new['BD2':'BU{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                for cell in item:  # cell表示每一行的每一个单元格对象
                    cell.fill = fillr
                    cell.border = Border(top=double, left=thin, right=thin, bottom=double)

            ws_new.merge_cells('BV1:CM1')
            merge_cell = ws_new['BV1']
            merge_cell.value = '平均CT值'
            merge_cell.alignment = Alignment(horizontal='center', vertical='center')
            merge_cell.fill = fille
            merge_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            for item in ws_new['BV2':'CM{}'.format(sheet.max_row + 1)]:  # item表示每一行的单元格元组
                for cell in item:  # cell表示每一行的每一个单元格对象
                    cell.fill = fille
                    cell.border = Border(top=double, left=thin, right=thin, bottom=double)

            # 字体设置
            for item in ws_new['A2':'CM{}'.format(sheet.max_row + 1)]:
                for cell in item:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for item in ws_new['A1':'CM2']:
                for cell in item:
                    cell.font = fonte
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            # 自适应列宽
            col_width = []
            i = 0
            # 每列
            for col in sheet.columns:
                # 每行
                for j in range(len(col)):
                    if j == 0:
                        # 数组增加一个元素
                        col_width.append(len(str(col[j].value)))
                    else:
                        # 获得每列中的内容的最大宽度
                        if col_width[i] < len(str(col[j].value)):
                            col_width[i] = len(str(col[j].value))
                i = i + 1

            # 设置列宽
            for i in range(len(col_width)):
                # 根据列的数字返回字母
                col_letter = get_column_letter(i + 1)
                # 当宽度大于100，宽度设置为100
                if col_width[i] > 100:
                    ws_new.column_dimensions[col_letter].width = 100
                # 只有当宽度大于10，才设置列宽
                elif col_width[i] > 1:
                    ws_new.column_dimensions[col_letter].width = col_width[i] + 2
        # 删除新建excel中多余的'sheet'页
        wb_new.remove_sheet(wb_new.get_sheet_by_name('Sheet'))
        # 保存新excel
        wb_new.save(new_xlsx_path)
        print('操作完成，请查看文件！')


if __name__ == '__main__':
    # n * sigma，设置倍数
    N = 1
    INPUT_FILE =r'D:\project\batch_processing_viscera-abnormal\腹部脏器统计_20200823.xlsx'
    # OUTPUT_PATH = r'C:\Users\SMIT\Desktop'
    va = VisveraAbnormal(INPUT_FILE, N)
    sheetnames = va.get_sheet()
    va.mjm()
    for i in sheetnames:
        if i == '门静脉':
            va.mjm()
            continue
        datas = va.build_dataframe(i)
        va.col_analyze(datas)
    va.operation_excel()
