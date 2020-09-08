import pandas as pd

# s1 = pd.Series(index=list("abcd"))
# print(s1)
# s1['a'] = 'asdf'
# print(s1.index[0])

# test_dict = {'i


# list1 = ['no.1', 'nox.2df', 'noc', 'noc.1df', 'noc.4df']
# for i in list1:
#     a = i.replace('.1', '')
#     print(a)

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors, Font, Alignment, PatternFill, Color, Side, Border
from openpyxl.utils import get_column_letter

wb_new = load_workbook(r'C:\Users\SMIT\Desktop\new.xlsx')
sheet = wb_new.active
wb_new.create_sheet('肝脏11')
wb_new.save(r'C:\Users\SMIT\Desktop\new.xlsx')


# 获取每一列的内容的最大宽度
# col_width = []
# i = 0
# # 每列
# for col in sheet.columns:
#     # 每行
#     for j in range(len(col)):
#         if j == 0:
#             # 数组增加一个元素
#             col_width.append(len(str(col[j].value)))
#         else:
#             # 获得每列中的内容的最大宽度
#             if col_width[i] < len(str(col[j].value)):
#                 col_width[i] = len(str(col[j].value))
#     i = i + 1
#
# # 设置列宽
# for i in range(len(col_width)):
#     # 根据列的数字返回字母
#     col_letter = get_column_letter(i + 1)
#     # 当宽度大于100，宽度设置为100
#     if col_width[i] > 100:
#         sheet.column_dimensions[col_letter].width = 100
#     # 只有当宽度大于10，才设置列宽
#     elif col_width[i] > 1:
#         sheet.column_dimensions[col_letter].width = col_width[i] + 2
# sheet.merge_cells('B1:M1')
# merge_cell = sheet['B1']
# merge_cell.value = '径线X'
# sheet.cell(row=1, column=2).fill = fille
# merge_cell.alignment = Alignment(horizontal='center', vertical='center')
# fille = PatternFill('solid', fgColor='C4E1FF')
# thin = Side(border_style="thin", color="000000")
# double = Side(border_style="thin", color="000000")
# for item in sheet['B2':'M{}'.format(sheet.max_row)]:  # item表示每一行的单元格元组
#     for cell in item:  # cell表示每一行的每一个单元格对象
#         cell.fill = fille
#         cell.border = Border(top=double, left=thin, right=thin, bottom=double)
# 注意保存，保存后excel文件才存在
